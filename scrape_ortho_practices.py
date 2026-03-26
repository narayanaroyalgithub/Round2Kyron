"""
Orthopedic Surgery Practice Dataset Builder (v2 - Fixed)
=========================================================
Uses the NPPES NPI Registry API to collect orthopedic surgery practices across the US.

KEY FIX from v1: The taxonomy_description parameter expects TEXT descriptions
(e.g., "Orthopaedic Surgery"), NOT taxonomy codes (e.g., "207X00000X").
Also uses the 'skip' parameter for proper pagination beyond 200 results.

Data source: NPPES NPI Registry (public, free, no API key required)
API: https://npiregistry.cms.hhs.gov/api/?version=2.1

Usage:
    pip install requests openpyxl pandas
    python scrape_ortho_practices.py
"""

import requests
import pandas as pd
import time
import json
import os
import sys
from collections import defaultdict
from datetime import datetime

# ============================================================
# CONFIGURATION
# ============================================================

API_BASE = "https://npiregistry.cms.hhs.gov/api/"
API_VERSION = "2.1"
RESULTS_PER_PAGE = 200      # API max per single call
MAX_RESULTS_TOTAL = 1200    # API hard cap per query combo
RATE_LIMIT_DELAY = 0.3      # seconds between API calls

# taxonomy_description expects TEXT, not codes.
# Wildcards (*) are supported after 2+ characters.
# "Orthopaedic*" will match all orthopaedic sub-specialties.
ORTHO_TAXONOMY_DESCRIPTIONS = [
    "Orthopaedic Surgery",                         # general ortho
    "Adult Reconstructive Orthopaedic Surgery",     # hip/knee replacement
    "Foot and Ankle Surgery",                       # foot & ankle
    "Hand Surgery",                                 # hand
    "Orthopaedic Surgery of the Spine",             # spine
    "Orthopaedic Trauma",                           # trauma
    "Pediatric Orthopaedic Surgery",                # pediatric
    "Sports Medicine",                              # sports med
]

US_STATES = [
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "DC", "FL",
    "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME",
    "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH",
    "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI",
    "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI",
    "WY", "PR", "VI", "GU", "AS", "MP"
]

OUTPUT_DIR = "output"
PROGRESS_FILE = "progress_v2.json"


# ============================================================
# API QUERY FUNCTIONS
# ============================================================

def query_nppes(params, retries=3):
    """Query the NPPES API with retry logic."""
    params["version"] = API_VERSION

    for attempt in range(retries):
        try:
            resp = requests.get(API_BASE, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()

            # Debug: print result count for first few calls
            result_count = data.get("result_count", 0)
            return data
        except requests.exceptions.RequestException as e:
            print(f"  [WARN] API request failed (attempt {attempt+1}/{retries}): {e}")
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
            else:
                print(f"  [ERROR] Giving up after {retries} attempts")
                return None
    return None


def search_paginated(params_base, max_total=MAX_RESULTS_TOTAL):
    """
    Paginate through NPPES API results using the 'skip' parameter.
    The API returns max 200 per call, and supports up to 1200 total per query.
    """
    all_results = []
    skip = 0

    while skip < max_total:
        params = {**params_base, "limit": RESULTS_PER_PAGE, "skip": skip}
        data = query_nppes(params)

        if not data or "results" not in data:
            break

        results = data["results"]
        result_count = data.get("result_count", 0)
        all_results.extend(results)

        # If we got fewer than 200, there are no more results
        if len(results) < RESULTS_PER_PAGE:
            break

        skip += RESULTS_PER_PAGE
        time.sleep(RATE_LIMIT_DELAY)

        # Safety: if result_count tells us we've got everything
        if len(all_results) >= result_count:
            break

    return all_results


# ============================================================
# DATA EXTRACTION
# ============================================================

def safe(val, default=""):
    """Return val if truthy, else default. Handles JSON null -> Python None."""
    return val if val is not None else default


def extract_practice_from_result(result):
    """Extract practice info from an API result (works for both NPI-1 and NPI-2)."""
    basic = result.get("basic", {})
    addresses = result.get("addresses", [])
    taxonomies = result.get("taxonomies", [])
    enumeration_type = result.get("enumeration_type", "")

    # Get practice location address (prefer LOCATION over MAILING)
    location_addr = None
    mailing_addr = None
    for addr in addresses:
        purpose = addr.get("address_purpose", "")
        if purpose == "LOCATION":
            location_addr = addr
        elif purpose == "MAILING":
            mailing_addr = addr

    addr = location_addr or mailing_addr or {}

    # Get taxonomy info
    primary_taxonomy = ""
    taxonomy_code = ""
    all_taxonomies = []
    for tax in taxonomies:
        desc = tax.get("desc") or ""
        all_taxonomies.append(desc)
        if tax.get("primary", False):
            primary_taxonomy = desc
            taxonomy_code = tax.get("code") or ""

    if not primary_taxonomy and all_taxonomies:
        primary_taxonomy = all_taxonomies[0]

    # Build the practice name
    if enumeration_type == "NPI-2":
        name = safe(basic.get("organization_name", basic.get("name")))
    else:
        name = f"{safe(basic.get('first_name'))} {safe(basic.get('last_name'))}".strip()
        credential = safe(basic.get("credential"))
        if credential:
            name = f"{name}, {credential}"

    practice = {
        "npi": safe(result.get("number")),
        "enumeration_type": safe(enumeration_type),
        "practice_name": safe(name),
        "phone": safe(addr.get("telephone_number")),
        "fax": safe(addr.get("fax_number")),
        "address_line_1": safe(addr.get("address_1")),
        "address_line_2": safe(addr.get("address_2")),
        "city": safe(addr.get("city")),
        "state": safe(addr.get("state")),
        "zip_code": safe(addr.get("postal_code")),
        "country": safe(addr.get("country_name")),
        "primary_taxonomy": safe(primary_taxonomy),
        "taxonomy_code": safe(taxonomy_code),
        "all_taxonomies": "; ".join(t for t in all_taxonomies if t),
        "enumeration_date": safe(basic.get("enumeration_date")),
        "last_updated": safe(basic.get("last_updated")),
        "status": safe(basic.get("status")),
        "authorized_official_name": f"{safe(basic.get('authorized_official_first_name'))} {safe(basic.get('authorized_official_last_name'))}".strip(),
        "authorized_official_title": safe(basic.get("authorized_official_title_or_position")),
    }
    return practice


# ============================================================
# PROGRESS / RESUME SUPPORT
# ============================================================

def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {"completed_keys": [], "practices": {}}


def save_progress(progress):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f)


# ============================================================
# MAIN COLLECTION LOGIC
# ============================================================

def collect_all_practices():
    """Main function: iterate over all states and taxonomy descriptions."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    progress = load_progress()

    all_practices = progress.get("practices", {})
    completed_keys = set(progress.get("completed_keys", []))

    total_api_calls = 0
    start_time = datetime.now()

    print("=" * 70)
    print("ORTHOPEDIC SURGERY PRACTICE DATASET BUILDER v2")
    print(f"Started at: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Taxonomy descriptions to search: {len(ORTHO_TAXONOMY_DESCRIPTIONS)}")
    print(f"States/territories to search: {len(US_STATES)}")
    print(f"Already completed: {len(completed_keys)} state/taxonomy combos")
    print(f"Already collected: {len(all_practices)} unique practices")
    print("=" * 70)

    # -------------------------------------------------------
    # PHASE 1: Organizations (NPI-2) -- these are practices
    # -------------------------------------------------------
    print("\n--- PHASE 1: Collecting Organization (NPI-2) NPIs ---\n")

    for tax_desc in ORTHO_TAXONOMY_DESCRIPTIONS:
        for state in US_STATES:
            key = f"org|{tax_desc}|{state}"
            if key in completed_keys:
                continue

            params = {
                "taxonomy_description": tax_desc,
                "enumeration_type": "NPI-2",
                "state": state,
            }

            results = search_paginated(params)
            new_count = 0

            for r in results:
                practice = extract_practice_from_result(r)
                npi = practice["npi"]
                if npi and npi not in all_practices:
                    all_practices[npi] = practice
                    new_count += 1

            if new_count > 0 or len(results) > 0:
                print(f"  [{state}] {tax_desc}: {len(results)} results, +{new_count} new (total: {len(all_practices)})")
            elif len(results) == 0:
                # silent for zero results to reduce noise
                pass

            completed_keys.add(key)
            time.sleep(RATE_LIMIT_DELAY)

        # Save progress after each taxonomy description
        progress["completed_keys"] = list(completed_keys)
        progress["practices"] = all_practices
        save_progress(progress)
        print(f"  >> Completed '{tax_desc}' -- total unique practices: {len(all_practices)}\n")

    # -------------------------------------------------------
    # PHASE 2: Individuals (NPI-1) for physician count enrichment
    # -------------------------------------------------------
    print("\n--- PHASE 2: Collecting Individual (NPI-1) NPIs for physician counts ---\n")

    address_to_physicians = defaultdict(list)

    # Only use the main taxonomy for individuals to save time
    ind_taxonomies = ["Orthopaedic Surgery"]

    for tax_desc in ind_taxonomies:
        for state in US_STATES:
            key = f"ind|{tax_desc}|{state}"
            if key in completed_keys:
                continue

            params = {
                "taxonomy_description": tax_desc,
                "enumeration_type": "NPI-1",
                "state": state,
            }

            results = search_paginated(params)

            for r in results:
                practice = extract_practice_from_result(r)
                addr_key = f"{practice['address_line_1']}|{practice['city']}|{practice['state']}".upper().strip()
                if addr_key and addr_key != "||":
                    address_to_physicians[addr_key].append(practice["practice_name"])

            if len(results) > 0:
                print(f"  [{state}] {len(results)} individual ortho providers found")

            completed_keys.add(key)
            time.sleep(RATE_LIMIT_DELAY)

        progress["completed_keys"] = list(completed_keys)
        save_progress(progress)

    # -------------------------------------------------------
    # PHASE 3: Enrich with physician counts
    # -------------------------------------------------------
    print("\n--- PHASE 3: Enriching practices with physician counts ---\n")

    for npi, practice in all_practices.items():
        addr_key = f"{practice['address_line_1']}|{practice['city']}|{practice['state']}".upper().strip()
        physicians = address_to_physicians.get(addr_key, [])
        practice["estimated_physician_count"] = len(physicians)
        practice["physician_names_sample"] = "; ".join(physicians[:10])

    # -------------------------------------------------------
    # PHASE 4: Export
    # -------------------------------------------------------
    print("--- PHASE 4: Exporting dataset ---\n")

    practices_list = list(all_practices.values())

    if not practices_list:
        print("  [ERROR] No practices collected! Something went wrong with the API.")
        print("  Try testing this URL in your browser:")
        print("  https://npiregistry.cms.hhs.gov/api/?version=2.1&taxonomy_description=Orthopaedic+Surgery&enumeration_type=NPI-2&state=CA&limit=5")
        return None

    df = pd.DataFrame(practices_list)

    # Filter out deactivated NPIs
    if "status" in df.columns:
        active_before = len(df)
        df = df[df["status"] != "D"]
        removed = active_before - len(df)
        if removed > 0:
            print(f"  Removed {removed} deactivated NPIs")

    # Track with and without phone
    if "phone" in df.columns:
        has_phone = df["phone"].notna() & (df["phone"] != "")
        no_phone_count = (~has_phone).sum()
        print(f"  Practices without phone numbers: {no_phone_count}")
        df_with_phone = df[has_phone].copy()
    else:
        print("  [WARN] No 'phone' column found. Exporting all records.")
        df_with_phone = df.copy()

    # Clean phone numbers
    if "phone" in df_with_phone.columns:
        df_with_phone["phone"] = df_with_phone["phone"].apply(format_phone)

    # Sort
    df_with_phone = df_with_phone.sort_values(["state", "city", "practice_name"])

    # Reorder columns
    column_order = [
        "practice_name", "phone", "fax",
        "address_line_1", "address_line_2", "city", "state", "zip_code",
        "primary_taxonomy", "taxonomy_code",
        "estimated_physician_count", "physician_names_sample",
        "npi", "enumeration_type", "enumeration_date", "last_updated",
        "authorized_official_name", "authorized_official_title",
        "all_taxonomies", "country"
    ]
    existing_cols = [c for c in column_order if c in df_with_phone.columns]
    df_with_phone = df_with_phone[existing_cols]

    # Save CSV
    csv_path = os.path.join(OUTPUT_DIR, "orthopedic_practices_dataset.csv")
    df_with_phone.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"  CSV saved: {csv_path}")

    # Save Excel
    xlsx_path = os.path.join(OUTPUT_DIR, "orthopedic_practices_dataset.xlsx")
    save_to_excel(df_with_phone, xlsx_path)
    print(f"  Excel saved: {xlsx_path}")

    # Also save the full dataset (including those without phone)
    csv_all_path = os.path.join(OUTPUT_DIR, "orthopedic_practices_ALL.csv")
    df.to_csv(csv_all_path, index=False, encoding="utf-8-sig")
    print(f"  Full CSV (incl. no phone): {csv_all_path}")

    # Summary
    elapsed = (datetime.now() - start_time).total_seconds()
    print("\n" + "=" * 70)
    print("COLLECTION COMPLETE")
    print(f"  Total unique practices (with phone): {len(df_with_phone)}")
    print(f"  Total unique practices (all):        {len(df)}")
    print(f"  Time elapsed:                        {elapsed/60:.1f} minutes")
    print(f"  Output files in:                     {OUTPUT_DIR}/")
    print("=" * 70)

    return df_with_phone


def format_phone(phone):
    """Format phone number to (XXX) XXX-XXXX."""
    if not phone:
        return ""
    digits = ''.join(c for c in str(phone) if c.isdigit())
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    elif len(digits) == 11 and digits[0] == '1':
        return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    return phone


def save_to_excel(df, filepath):
    """Save DataFrame to a nicely formatted Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Orthopedic Practices"

        header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        data_font = Font(name="Arial", size=10)

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value if pd.notna(value) else "")
                cell.border = thin_border
                cell.font = data_font
                cell.alignment = Alignment(vertical="center")

        # Auto-fit column widths
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max() if len(df) > 0 else 0
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wb.save(filepath)
    except ImportError:
        print("  [WARN] openpyxl not installed. Saving as CSV only.")
        df.to_csv(filepath.replace(".xlsx", "_fallback.csv"), index=False)


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":
    print("\nChecking dependencies...")
    try:
        import requests
        print("  requests: OK")
    except ImportError:
        print("  [MISSING] requests -- pip install requests")
        sys.exit(1)
    try:
        import pandas
        print("  pandas: OK")
    except ImportError:
        print("  [MISSING] pandas -- pip install pandas")
        sys.exit(1)
    try:
        import openpyxl
        print("  openpyxl: OK")
    except ImportError:
        print("  openpyxl: NOT FOUND (will save CSV only, install with: pip install openpyxl)")

    # Quick sanity test: make one API call to verify connectivity
    print("\nTesting API connectivity...")
    test_params = {
        "version": API_VERSION,
        "taxonomy_description": "Orthopaedic Surgery",
        "enumeration_type": "NPI-2",
        "state": "CA",
        "limit": 3,
    }
    try:
        test_resp = requests.get(API_BASE, params=test_params, timeout=15)
        test_data = test_resp.json()
        test_count = test_data.get("result_count", 0)
        print(f"  API test: {test_count} results for Orthopaedic Surgery orgs in CA")
        if test_count == 0:
            print("  [WARN] Got 0 results. The API might be down or parameters may need adjustment.")
            print(f"  Test URL: {test_resp.url}")
            print(f"  Response: {json.dumps(test_data, indent=2)[:500]}")
            
            # Try alternative: use wildcard
            print("\n  Trying wildcard search: 'Orthopaedic*'...")
            test_params2 = {
                "version": API_VERSION,
                "taxonomy_description": "Orthopaedic*",
                "state": "CA",
                "limit": 3,
            }
            test_resp2 = requests.get(API_BASE, params=test_params2, timeout=15)
            test_data2 = test_resp2.json()
            test_count2 = test_data2.get("result_count", 0)
            print(f"  Wildcard test: {test_count2} results")
            if test_count2 > 0:
                print("  Wildcard works! Switching to wildcard mode.")
                # Replace the taxonomy list with wildcard
                ORTHO_TAXONOMY_DESCRIPTIONS.clear()
                ORTHO_TAXONOMY_DESCRIPTIONS.append("Orthopaedic*")

            # Also try without enumeration_type filter
            print("\n  Trying without enumeration_type filter...")
            test_params3 = {
                "version": API_VERSION,
                "taxonomy_description": "Orthopaedic Surgery",
                "state": "CA",
                "limit": 3,
            }
            test_resp3 = requests.get(API_BASE, params=test_params3, timeout=15)
            test_data3 = test_resp3.json()
            test_count3 = test_data3.get("result_count", 0)
            print(f"  No enum filter test: {test_count3} results")
            if test_count3 > 0:
                print(f"  Response sample: {json.dumps(test_data3.get('results', [{}])[0].get('basic', {}), indent=2)[:300]}")
        else:
            print("  API is working! Starting full collection...\n")
            # Print a sample result to verify structure
            if test_data.get("results"):
                sample = test_data["results"][0]
                print(f"  Sample: {sample.get('basic', {}).get('organization_name', 'N/A')} "
                      f"in {sample.get('addresses', [{}])[0].get('city', 'N/A')}, "
                      f"{sample.get('addresses', [{}])[0].get('state', 'N/A')}")
    except Exception as e:
        print(f"  [ERROR] API test failed: {e}")
        print("  Check your internet connection and try again.")
        sys.exit(1)

    print()
    collect_all_practices()